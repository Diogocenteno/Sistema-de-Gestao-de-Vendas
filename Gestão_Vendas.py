#%%


# -*- coding: utf-8 -*-
# Gestão de Vendas com Tkinter, SQLite e Pandas
# Autor: Mutante
# Data: 02/10/2025
# Versão: 2.34 (Permite escolher o local de salvamento na exportação)
# Descrição: Aplicação para gerenciar vendas, encomendas, anotações e visualizar análises de dados.

# Certifique-se de ter as bibliotecas instaladas:
# pip install ttkbootstrap pandas openpyxl matplotlib

# --- Importações Unificadas ---
import sqlite3
import pandas as pd
import tkinter as tk
from tkinter import messagebox, ttk, END, BOTH, YES, X, LEFT, HORIZONTAL, filedialog
import os
import sys
from ttkbootstrap import Style
import ttkbootstrap as ttk_bootstrap
from ttkbootstrap.constants import *
from ttkbootstrap.dialogs import Messagebox
from ttkbootstrap.tooltip import ToolTip
from ttkbootstrap.scrolled import ScrolledFrame # Usado na Calculadora
from datetime import datetime
import subprocess
import ctypes
import openpyxl
import threading
import json # Usado na Calculadora

# Importações para a nova aba de Dashboard
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt


# --- Função para encontrar o caminho dos arquivos (essencial para o PyInstaller) ---
def resource_path(relative_path):
    """ Retorna o caminho absoluto para o recurso, funciona para dev e para PyInstaller """
    try:
        # PyInstaller cria uma pasta temporária e armazena o caminho em _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


# --- Classe para Gerenciamento do Banco de Dados ---
class DatabaseManager:
    """
    Gerencia a conexão e as operações com o banco de dados SQLite.
    Responsável por inicializar o DB e todas as operações CRUD para
    vendas, encomendas e anotações.
    """
    def __init__(self, db_path):
        self.db_path = db_path
        self.conn = None
        self.cursor = None
        self._initialize_db()

    def _initialize_db(self):
        """
        Inicializa a conexão com o banco de dados e cria as tabelas se não existirem.
        Oculta o arquivo do banco de dados no Windows.
        """
        try:
            self.conn = sqlite3.connect(self.db_path, check_same_thread=False) # Habilita para threads
            self.cursor = self.conn.cursor()
            
            # 1. Cria a tabela 'vendas'
            self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS vendas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome_cliente TEXT,
                nome_produto TEXT,
                quantidade INTEGER,
                preco REAL,
                tipo_pagamento TEXT,
                preco_final REAL,
                nome_vendedor TEXT,
                data_hora TEXT
            )
            """)

            # 2. Cria a tabela 'encomendas'
            self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS encomendas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                data_hora_registro TEXT,
                nome_cliente TEXT,
                produto TEXT,
                quantidade INTEGER,
                valor_unitario REAL,
                data_entrega TEXT
            )
            """)

            # 3. Cria a tabela 'anotacoes'
            self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS anotacoes (
                id INTEGER PRIMARY KEY DEFAULT 1, -- Sempre usará o ID 1 para ter um único registro
                conteudo TEXT
            )
            """)
            # Garante que a linha de anotações exista
            self.cursor.execute("INSERT OR IGNORE INTO anotacoes (id, conteudo) VALUES (1, '')")

            self.conn.commit()

            # Ocultar o arquivo do banco de dados (específico para Windows)
            if os.name == 'nt' and os.path.exists(self.db_path):
                try:
                    ctypes.windll.kernel32.SetFileAttributesW(self.db_path, 0x02) # Atributo HIDDEN
                except Exception as e:
                    print(f"Não foi possível ocultar o arquivo do banco de dados: {e}")

        except sqlite3.Error as e:
            messagebox.showerror("Erro de Banco de Dados", f"Erro ao inicializar: {e}")
            exit()

    # --- Métodos para Vendas ---
    def insert_sale(self, data):
        """Insere uma nova venda no banco de dados."""
        try:
            self.cursor.execute("""
            INSERT INTO vendas (nome_cliente, nome_produto, quantidade, preco, tipo_pagamento, preco_final, nome_vendedor, data_hora)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (data["nome_cliente"], data["nome_produto"], data["quantidade"], data["preco"],
                  data["tipo_pagamento"], data["preco_final"], data["nome_vendedor"], data["data_hora"]))
            self.conn.commit()
        except sqlite3.Error as e:
            self.conn.rollback()
            messagebox.showerror("Erro de Banco de Dados", f"Erro ao inserir venda: {e}")
    
    def insert_multiple_sales(self, sales_data_list):
        """Insere múltiplas vendas em uma única transação."""
        try:
            self.cursor.executemany("""
            INSERT INTO vendas (nome_cliente, nome_produto, quantidade, preco, tipo_pagamento, preco_final, nome_vendedor, data_hora)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, sales_data_list)
            self.conn.commit()
            return True
        except sqlite3.Error as e:
            self.conn.rollback()
            messagebox.showerror("Erro de Banco de Dados", f"Erro ao inserir múltiplas vendas: {e}")
            return False

    def update_sale(self, sale_id, data):
        """Atualiza uma venda existente."""
        try:
            self.cursor.execute("""
            UPDATE vendas
            SET nome_cliente=?, nome_produto=?, quantidade=?, preco=?, tipo_pagamento=?, preco_final=?, nome_vendedor=?, data_hora=?
            WHERE id=?
            """, (data["nome_cliente"], data["nome_produto"], data["quantidade"], data["preco"],
                  data["tipo_pagamento"], data["preco_final"], data["nome_vendedor"], data["data_hora"], sale_id))
            self.conn.commit()
        except sqlite3.Error as e:
            self.conn.rollback()
            messagebox.showerror("Erro de Banco de Dados", f"Erro ao atualizar venda: {e}")

    def delete_sale(self, sale_id):
        """Exclui uma venda pelo ID."""
        try:
            self.cursor.execute("DELETE FROM vendas WHERE id=?", (sale_id,))
            self.conn.commit()
        except sqlite3.Error as e:
            self.conn.rollback()
            messagebox.showerror("Erro de Banco de Dados", f"Erro ao excluir venda: {e}")

    def fetch_all_sales(self, search_term=""):
        """Busca todas as vendas, opcionalmente filtrando por um termo de busca."""
        try:
            query = "SELECT id, data_hora, nome_cliente, nome_produto, quantidade, preco, tipo_pagamento, preco_final, nome_vendedor FROM vendas"
            params = []
            if search_term:
                search_pattern = f"%{search_term}%"
                query += " WHERE nome_cliente LIKE ? OR nome_produto LIKE ? OR nome_vendedor LIKE ?"
                params.extend([search_pattern, search_pattern, search_pattern])
            query += " ORDER BY id DESC"
            self.cursor.execute(query, params)
            return self.cursor.fetchall()
        except sqlite3.Error as e:
            messagebox.showerror("Erro de Banco de Dados", f"Erro ao buscar vendas: {e}")
            return []
    
    def fetch_sales_as_dataframe(self):
        """Busca todas as vendas e retorna como um DataFrame do Pandas."""
        try:
            df = pd.read_sql_query("SELECT * FROM vendas", self.conn)
            # Converte a coluna de data para o formato datetime do pandas
            df['data_hora'] = pd.to_datetime(df['data_hora'], format='%d/%m/%Y %H:%M:%S', errors='coerce')
            return df
        except Exception as e:
            messagebox.showerror("Erro de Banco de Dados", f"Erro ao buscar vendas para análise: {e}")
            return pd.DataFrame() # Retorna um DataFrame vazio em caso de erro

    def fetch_sale_by_id(self, sale_id):
        """Busca uma venda específica pelo ID."""
        try:
            self.cursor.execute("SELECT id, nome_cliente, nome_produto, quantidade, preco, tipo_pagamento, nome_vendedor FROM vendas WHERE id=?", (sale_id,))
            return self.cursor.fetchone()
        except sqlite3.Error as e:
            messagebox.showerror("Erro de Banco de Dados", f"Erro ao buscar venda por ID: {e}")
            return None

    # --- Métodos para Encomendas ---
    def insert_encomenda(self, data):
        """Insere uma nova encomenda."""
        try:
            self.cursor.execute("""
            INSERT INTO encomendas (data_hora_registro, nome_cliente, produto, quantidade, valor_unitario, data_entrega)
            VALUES (?, ?, ?, ?, ?, ?)
            """, (data["data_hora_registro"], data["nome_cliente"], data["produto"],
                  data["quantidade"], data["valor_unitario"], data["data_entrega"]))
            self.conn.commit()
        except sqlite3.Error as e:
            self.conn.rollback()
            messagebox.showerror("Erro de Banco de Dados", f"Erro ao inserir encomenda: {e}")

    def update_encomenda(self, encomenda_id, data):
        """Atualiza uma encomenda."""
        try:
            self.cursor.execute("""
            UPDATE encomendas SET nome_cliente=?, produto=?, quantidade=?, valor_unitario=?, data_entrega=?
            WHERE id=?
            """, (data["nome_cliente"], data["produto"], data["quantidade"],
                  data["valor_unitario"], data["data_entrega"], encomenda_id))
            self.conn.commit()
        except sqlite3.Error as e:
            self.conn.rollback()
            messagebox.showerror("Erro de Banco de Dados", f"Erro ao atualizar encomenda: {e}")

    def delete_encomenda(self, encomenda_id):
        """Exclui uma encomenda."""
        try:
            self.cursor.execute("DELETE FROM encomendas WHERE id=?", (encomenda_id,))
            self.conn.commit()
        except sqlite3.Error as e:
            self.conn.rollback()
            messagebox.showerror("Erro de Banco de Dados", f"Erro ao excluir encomenda: {e}")
    
    def clear_all_encomendas(self):
        """Limpa todas as encomendas da tabela."""
        try:
            self.cursor.execute("DELETE FROM encomendas")
            self.conn.commit()
        except sqlite3.Error as e:
            self.conn.rollback()
            messagebox.showerror("Erro de Banco de Dados", f"Erro ao limpar encomendas: {e}")

    def fetch_all_encomendas(self):
        """Busca todas as encomendas."""
        try:
            self.cursor.execute("SELECT id, data_hora_registro, nome_cliente, produto, quantidade, valor_unitario, data_entrega FROM encomendas ORDER BY id DESC")
            return self.cursor.fetchall()
        except sqlite3.Error as e:
            messagebox.showerror("Erro de Banco de Dados", f"Erro ao buscar encomendas: {e}")
            return []

    # --- Métodos para Anotações ---
    def fetch_anotacoes(self):
        """Busca o conteúdo das anotações."""
        try:
            self.cursor.execute("SELECT conteudo FROM anotacoes WHERE id=1")
            result = self.cursor.fetchone()
            return result[0] if result else ""
        except sqlite3.Error as e:
            messagebox.showerror("Erro de Banco de Dados", f"Erro ao buscar anotações: {e}")
            return ""

    def save_anotacoes(self, content):
        """Salva o conteúdo das anotações."""
        try:
            # UPDATE OR INSERT para garantir que o registro exista
            self.cursor.execute("UPDATE anotacoes SET conteudo=? WHERE id=1", (content,))
            self.conn.commit()
        except sqlite3.Error as e:
            self.conn.rollback()
            messagebox.showerror("Erro de Banco de Dados", f"Erro ao salvar anotações: {e}")

    def close_connection(self):
        """Fecha a conexão com o banco de dados e reexibe o arquivo no Windows."""
        if self.conn:
            self.conn.close()
        if os.name == 'nt' and os.path.exists(self.db_path):
            try:
                # Atributo NORMAL
                ctypes.windll.kernel32.SetFileAttributesW(self.db_path, 0x80)
            except Exception as e:
                print(f"Não foi possível reexibir o arquivo do banco de dados: {e}")


# --- Classe para o Caderno Virtual de Encomendas ---
class CadernoVirtual:
    """
    Gerencia a janela e a lógica para as encomendas,
    agora utilizando o banco de dados.
    """
    TOTAL_ROW_ID = "caderno_total_row"

    def __init__(self, parent_root, theme_style, db_manager):
        self.db_manager = db_manager
        self.caderno_window = tk.Toplevel(parent_root)
        self.caderno_window.title("Caderno de Encomendas")
        self.caderno_window.geometry("1100x750")
        self.caderno_window.transient(parent_root)
        self.caderno_window.grab_set()

        self.style = theme_style
        self._setup_styles()

        self._create_widgets()
        self._load_content()
        self._reset_input_fields()
    
    def _setup_styles(self):
        self.style.configure('Caderno.TFrame', background=self.style.lookup('TFrame', 'background'))
        self.style.configure('Caderno.TButton', font=('Arial', 10, 'bold'))
        self.style.configure('Caderno.TLabel', font=('Arial', 12, 'bold'))
        self.style.configure('Caderno.Treeview.Heading', font=('Arial', 10, 'bold'))
        self.style.configure('Caderno.Treeview', rowheight=25)
        self.style.map('Caderno.Treeview', background=[('selected', self.style.colors.primary)])

    def _create_widgets(self):
        frame_caderno = ttk.Frame(self.caderno_window, padding=10, style='Caderno.TFrame')
        frame_caderno.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame_caderno, text="Suas Encomendas", font=('Arial', 16, 'bold'), style='Caderno.TLabel').pack(pady=(0, 10))

        frame_input = ttk.LabelFrame(frame_caderno, text="Adicionar/Editar Encomenda", padding=10)
        frame_input.pack(fill=tk.X, pady=(0, 10))

        # Campos de entrada
        labels = ["Nome:", "Produto:", "Quantidade:", "Valor por unidade (R$):", "Data de Entrega:"]
        self.entries = {}
        for i, text in enumerate(labels):
            ttk.Label(frame_input, text=text).grid(row=i, column=0, sticky="w", padx=5, pady=2)
            entry = ttk.Entry(frame_input, width=30)
            entry.grid(row=i, column=1, sticky="ew", padx=5, pady=2)
            self.entries[text] = entry
            
        frame_input.grid_columnconfigure(1, weight=1)

        # Botões de ação
        frame_botoes_input = ttk.Frame(frame_input)
        frame_botoes_input.grid(row=len(labels), column=0, columnspan=2, pady=10, sticky="ew")

        self.btn_add_encomenda = ttk.Button(frame_botoes_input, text="Adicionar Encomenda", command=self._add_encomenda, style='success.TButton')
        self.btn_add_encomenda.pack(side=tk.LEFT, expand=True, padx=5, fill=tk.X)

        self.btn_update_encomenda = ttk.Button(frame_botoes_input, text="Atualizar Encomenda", command=self._update_encomenda, style='primary.TButton')
        self.btn_cancel_edit = ttk.Button(frame_botoes_input, text="Cancelar Edição", command=self._reset_input_fields, style='secondary.TButton')

        # Treeview para exibir as encomendas
        cols = ("ID", "Data e Hora", "Nome", "Produto", "Quantidade", "Valor por unidade", "Data Entrega", "Valor Total")
        self.tree = ttk.Treeview(frame_caderno, columns=cols, show="headings", style='Caderno.Treeview')
        self.tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Configuração das colunas
        headings = {"ID": 50, "Data e Hora": 150, "Nome": 150, "Produto": 180, "Quantidade": 80,
                    "Valor por unidade": 120, "Data Entrega": 120, "Valor Total": 110}
        
        for col, width in headings.items():
            self.tree.heading(col, text=f"{col} (R$)" if "Valor" in col else col)
            self.tree.column(col, width=width, anchor="center", stretch=tk.NO if col in ["ID", "Quantidade"] else tk.YES)

        # Scrollbars
        scrollbar_y = ttk.Scrollbar(self.tree, orient="vertical", command=self.tree.yview)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=scrollbar_y.set)
        
        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)
        self.selected_item_iid = None

        # Botões de gestão
        frame_botoes_gestao = ttk.Frame(frame_caderno, style='Caderno.TFrame')
        frame_botoes_gestao.pack(fill=tk.X, pady=(10, 0))

        ttk.Button(frame_botoes_gestao, text="Excluir Selecionada(s)", command=self._delete_encomenda, style='danger.TButton').pack(side=tk.LEFT, expand=True, padx=5)
        ttk.Button(frame_botoes_gestao, text="Limpar Caderno Completo", command=self._clear_all, style='warning.TButton').pack(side=tk.LEFT, expand=True, padx=5)
        ttk.Button(frame_botoes_gestao, text="Fechar Caderno", command=self.caderno_window.destroy, style='secondary.TButton').pack(side=tk.RIGHT, expand=True, padx=5)

    def _load_content(self):
        """Carrega encomendas do banco de dados para a Treeview."""
        self.tree.delete(*self.tree.get_children())
        encomendas = self.db_manager.fetch_all_encomendas()
        for enc in encomendas:
            encomenda_id, data_reg, nome, prod, qtd, val_unit, data_ent = enc
            try:
                valor_total = int(qtd) * float(val_unit)
                self.tree.insert('', tk.END, iid=encomenda_id, values=(
                    encomenda_id, data_reg, nome, prod, qtd,
                    f"{val_unit:.2f}".replace(".", ","),
                    data_ent,
                    f"{valor_total:.2f}".replace(".", ",")
                ))
            except (ValueError, TypeError):
                continue
        self._calculate_total()
    
    def _validate_inputs(self):
        """Valida e retorna os dados dos campos de entrada."""
        nome = self.entries["Nome:"].get().strip()
        produto = self.entries["Produto:"].get().strip()
        quantidade_str = self.entries["Quantidade:"].get().strip()
        valor_unitario_str = self.entries["Valor por unidade (R$):"].get().strip()
        data_entrega = self.entries["Data de Entrega:"].get().strip()

        if not all([nome, produto, quantidade_str, valor_unitario_str, data_entrega]):
            messagebox.showwarning("Atenção", "Todos os campos são obrigatórios.")
            return None

        try:
            quantidade = int(quantidade_str)
            if quantidade <= 0: raise ValueError
        except ValueError:
            messagebox.showwarning("Atenção", "Quantidade deve ser um número inteiro positivo.")
            return None

        try:
            valor_unitario = float(valor_unitario_str.replace(",", "."))
            if valor_unitario < 0: raise ValueError
        except ValueError:
            messagebox.showwarning("Atenção", "Valor por unidade inválido.")
            return None
            
        return {
            "nome_cliente": nome, "produto": produto, "quantidade": quantidade,
            "valor_unitario": valor_unitario, "data_entrega": data_entrega
        }

    def _add_encomenda(self):
        """Adiciona uma nova encomenda ao banco de dados."""
        data = self._validate_inputs()
        if data:
            data["data_hora_registro"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            self.db_manager.insert_encomenda(data)
            self._load_content()
            self._reset_input_fields()
    
    def _update_encomenda(self):
        """Atualiza a encomenda selecionada."""
        if not self.selected_item_iid:
            messagebox.showwarning("Atenção", "Nenhuma encomenda selecionada.")
            return

        data = self._validate_inputs()
        if data:
            self.db_manager.update_encomenda(self.selected_item_iid, data)
            messagebox.showinfo("Sucesso", "Encomenda atualizada!")
            self._load_content()
            self._reset_input_fields()

    def _on_tree_select(self, event):
        """Carrega os dados da encomenda selecionada nos campos de entrada."""
        selected_items = self.tree.selection()
        if not selected_items or selected_items[0] == self.TOTAL_ROW_ID:
            self._reset_input_fields()
            return

        self.selected_item_iid = selected_items[0]
        values = self.tree.item(self.selected_item_iid, 'values')
        
        self.entries["Nome:"].delete(0, tk.END); self.entries["Nome:"].insert(0, values[2])
        self.entries["Produto:"].delete(0, tk.END); self.entries["Produto:"].insert(0, values[3])
        self.entries["Quantidade:"].delete(0, tk.END); self.entries["Quantidade:"].insert(0, values[4])
        self.entries["Valor por unidade (R$):"].delete(0, tk.END); self.entries["Valor por unidade (R$):"].insert(0, values[5])
        self.entries["Data de Entrega:"].delete(0, tk.END); self.entries["Data de Entrega:"].insert(0, values[6])

        self.btn_add_encomenda.pack_forget()
        self.btn_update_encomenda.pack(side=tk.LEFT, expand=True, padx=5, fill=tk.X)
        self.btn_cancel_edit.pack(side=tk.LEFT, expand=True, padx=5, fill=tk.X)

    def _delete_encomenda(self):
        """Exclui a(s) encomenda(s) selecionada(s)."""
        selected_items = [item for item in self.tree.selection() if item != self.TOTAL_ROW_ID]
        if not selected_items:
            messagebox.showwarning("Atenção", "Selecione uma ou mais encomendas para excluir.")
            return

        msg = f"Tem certeza que deseja excluir {len(selected_items)} encomenda(s)?"
        if messagebox.askyesno("Confirmar Exclusão", msg):
            for item_id in selected_items:
                self.db_manager.delete_encomenda(item_id)
            messagebox.showinfo("Sucesso", f"{len(selected_items)} encomenda(s) excluída(s).")
            self._load_content()
            self._reset_input_fields()

    def _reset_input_fields(self):
        """Limpa campos de entrada e redefine botões."""
        for entry in self.entries.values():
            entry.delete(0, tk.END)
        
        self.selected_item_iid = None
        self.btn_update_encomenda.pack_forget()
        self.btn_cancel_edit.pack_forget()
        self.btn_add_encomenda.pack(side=tk.LEFT, expand=True, padx=5, fill=tk.X)
        if self.tree.selection():
            self.tree.selection_remove(self.tree.selection())

    def _clear_all(self):
        """Limpa todo o conteúdo do caderno."""
        if messagebox.askyesno("Confirmar", "Tem certeza que deseja limpar TODO o caderno?"):
            self.db_manager.clear_all_encomendas()
            messagebox.showinfo("Sucesso", "O caderno foi limpo.")
            self._load_content()
            self._reset_input_fields()

    def _calculate_total(self):
        """Calcula e exibe o total na Treeview."""
        if self.tree.exists(self.TOTAL_ROW_ID):
            self.tree.delete(self.TOTAL_ROW_ID)

        total_valor = sum(float(self.tree.item(item, 'values')[7].replace(",", "."))
                              for item in self.tree.get_children()
                              if item != self.TOTAL_ROW_ID)

        total_formatado = f"{total_valor:.2f}".replace(".", ",")
        self.tree.insert('', tk.END, iid=self.TOTAL_ROW_ID, values=(
            "", "", "", "", "", "", "TOTAL GERAL:", total_formatado
        ), tags=('total_row',))
        
        self.tree.tag_configure('total_row', background=self.style.lookup('TFrame', 'background'), font=('Arial', 10, 'bold'))


# --- Classe para as Anotações Virtuais ---
class AnotacoesVirtual:
    """Janela de anotações, utilizando o banco de dados."""
    def __init__(self, parent_root, theme_style, db_manager):
        self.db_manager = db_manager
        self.anotacoes_window = tk.Toplevel(parent_root)
        self.anotacoes_window.title("Anotações")
        self.anotacoes_window.geometry("500x600")
        self.anotacoes_window.transient(parent_root)
        self.anotacoes_window.grab_set()
        self.anotacoes_window.protocol("WM_DELETE_WINDOW", self._on_closing)

        self.style = theme_style
        self._create_widgets()
        self._load_content()
    
    def _create_widgets(self):
        frame_anotacoes = ttk.Frame(self.anotacoes_window, padding=10)
        frame_anotacoes.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame_anotacoes, text="Suas Anotações", font=('Arial', 18, 'bold')).pack(pady=(0, 10))

        self.text_area = tk.Text(frame_anotacoes, wrap=tk.WORD, font=('Arial', 12), relief=tk.SOLID, bd=1)
        self.text_area.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        frame_botoes = ttk.Frame(frame_anotacoes)
        frame_botoes.pack(fill=tk.X, pady=(10, 0))
        ttk.Button(frame_botoes, text="Salvar Anotações", command=self._save_content, style='info.TButton').pack(side=tk.LEFT, expand=True, padx=5)
        ttk.Button(frame_botoes, text="Limpar", command=self._clear_content, style='danger.TButton').pack(side=tk.LEFT, expand=True, padx=5)
        ttk.Button(frame_botoes, text="Fechar", command=self._on_closing, style='secondary.TButton').pack(side=tk.RIGHT, expand=True, padx=5)

    def _load_content(self):
        """Carrega anotações do banco de dados."""
        content = self.db_manager.fetch_anotacoes()
        self.text_area.delete(1.0, tk.END)
        self.text_area.insert(1.0, content)

    def _save_content(self):
        """Salva o conteúdo no banco de dados."""
        content = self.text_area.get(1.0, tk.END).strip()
        self.db_manager.save_anotacoes(content)
        messagebox.showinfo("Salvo", "Anotações salvas com sucesso!", parent=self.anotacoes_window)

    def _clear_content(self):
        """Limpa o campo de texto e salva."""
        if messagebox.askyesno("Confirmar", "Limpar todas as anotações?", parent=self.anotacoes_window):
            self.text_area.delete(1.0, tk.END)
            self._save_content()

    def _on_closing(self):
        """Salva ao fechar e destrói a janela."""
        self._save_content()
        self.anotacoes_window.destroy()


# --- ###################################################################### ---
# --- ### CLASSE DA CALCULADORA DE PREÇO ATUALIZADA E INTEGRADA ABAIXO ### ---
# --- ###################################################################### ---
class CalculadoraPrecoVenda(ttk.Frame):
    """
    Uma aplicação de calculadora de preço de venda ideal, adaptada para ser um Frame
    dentro de uma aba, com todas as melhorias recentes.
    """
    def __init__(self, parent, style):
        super().__init__(parent)
        self.style = style

        # --- Inicializa dicionários para guardar os dados ---
        self.inputs = {}
        self.results = {}

        # --- Validação de Entrada ---
        self.vcmd = (self.register(self.validate_input), '%P')

        # --- Container com rolagem ---
        scrolled_frame = ScrolledFrame(self, autohide=True)
        scrolled_frame.pack(fill=BOTH, expand=YES, padx=10, pady=10)

        # --- Entradas de Dados ---
        input_frame = ttk.LabelFrame(scrolled_frame, text=" 1. Insira os Dados do Produto ", padding="15")
        input_frame.pack(fill=X, pady=5)
        
        self.entry_nome_produto = self.create_entry_field(
            input_frame, "Nome do Produto:", "Nome ou descrição do item para salvar no histórico.", validate=False
        )
        self.entry_custo_produto = self.create_entry_field(
            input_frame, "Custo do Produto (R$):", "Custo de aquisição ou produção do item."
        )
        self.entry_gasto_operacional = self.create_entry_field(
            input_frame, "Gasto Operacional (R$):", "Custos fixos e variáveis diluídos por unidade."
        )
        self.entry_impostos_pct = self.create_entry_field(
            input_frame, "Alíquota de Imposto (%):", "Percentual de imposto sobre a venda."
        )
        self.entry_custo_transacao = self.create_entry_field(
            input_frame, "Taxa da Transação (%):", "Taxa da maquininha de cartão (opcional)."
        )
        self.entry_margem_lucro = self.create_entry_field(
            input_frame, "Lucro Desejado (%):", "Percentual de lucro sobre o preço de venda."
        )
        self.entry_unidades = self.create_entry_field(
            input_frame, "Dividir por Unidades:", "Opcional. Divida o resultado por um número de unidades."
        )

        # --- Botões de Ação ---
        action_frame = ttk.Frame(scrolled_frame)
        action_frame.pack(fill=X, pady=15)
        
        self.btn_calcular = ttk.Button(action_frame, text="Calcular Preço Ideal", command=self.calcular_preco, bootstyle="success")
        self.btn_calcular.pack(side=LEFT, fill=X, expand=YES, padx=(0, 5), ipady=5)

        self.btn_salvar = ttk.Button(action_frame, text="Salvar em Excel", command=self.salvar_excel, bootstyle="info")
        self.btn_salvar.pack(side=LEFT, fill=X, expand=YES, padx=5, ipady=5)
        
        self.btn_reset = ttk.Button(action_frame, text="Resetar Campos", command=self.resetar_campos, bootstyle="warning")
        self.btn_reset.pack(side=LEFT, fill=X, expand=YES, padx=(5, 0), ipady=5)

        # --- Resultados ---
        result_frame = ttk.LabelFrame(scrolled_frame, text=" 2. Resultados do Cálculo ", padding="15")
        result_frame.pack(fill=X, pady=5)
        
        self.preco_ideal_var = tk.StringVar(value="R$ 0,00")
        self.lucro_estimado_var = tk.StringVar(value="R$ 0,00")

        self.create_result_line(result_frame, "Lucro Bruto Total:", self.lucro_estimado_var, self.copy_to_clipboard, ("Helvetica", 12), bootstyle="info")
        self.create_result_line(result_frame, "Preço de Venda Total:", self.preco_ideal_var, self.copy_to_clipboard, ("Helvetica", 24, "bold"), bootstyle="success", pady_top=15)

        # --- Resultados por Unidade (inicialmente escondido) ---
        self.unit_result_frame = ttk.LabelFrame(scrolled_frame, text=" Resultados por Unidade ", padding="15")
        
        self.preco_unitario_var = tk.StringVar(value="R$ 0,00")
        self.lucro_unitario_var = tk.StringVar(value="R$ 0,00")

        self.create_result_line(self.unit_result_frame, "Lucro por Unidade:", self.lucro_unitario_var, self.copy_to_clipboard, ("Helvetica", 10), bootstyle="secondary")
        self.create_result_line(self.unit_result_frame, "Preço por Unidade:", self.preco_unitario_var, self.copy_to_clipboard, ("Helvetica", 16, "bold"), bootstyle="primary", pady_top=10)

        # --- Detalhamento do Preço (inicialmente vazio) ---
        self.details_frame = ttk.LabelFrame(scrolled_frame, text=" 3. Detalhamento do Preço de Venda ", padding="15")

    def _get_input_data(self):
        """Coleta e processa os dados de todos os campos de entrada."""
        return {
            'nome_produto': self.entry_nome_produto.get() or "Não informado",
            'custo_produto': self.parse_float(self.entry_custo_produto.get()),
            'gasto_operacional': self.parse_float(self.entry_gasto_operacional.get()),
            'impostos_pct': self.parse_float(self.entry_impostos_pct.get()),
            'custo_transacao_pct': self.parse_float(self.entry_custo_transacao.get()),
            'margem_lucro_pct': self.parse_float(self.entry_margem_lucro.get()),
            'unidades': self.parse_float(self.entry_unidades.get())
        }

    def create_result_line(self, parent, text, var, command, font, bootstyle, pady_top=0):
        frame = ttk.Frame(parent)
        frame.pack(fill=X, pady=(pady_top, 5))
        ttk.Label(frame, text=text, font=font).pack(side=LEFT, anchor="w")
        value_label = ttk.Label(frame, textvariable=var, font=font, bootstyle=bootstyle)
        value_label.pack(side=LEFT, padx=10, anchor="w")
        copy_button = ttk.Button(frame, text="📋", bootstyle="light-outline", command=lambda: command(var.get()))
        copy_button.pack(side=LEFT, padx=5, anchor="w")
        ToolTip(copy_button, "Copiar valor")

    def create_entry_field(self, parent, label_text, tooltip_text, validate=True):
        frame = ttk.Frame(parent)
        frame.pack(fill=X, pady=5)
        label = ttk.Label(frame, text=label_text, width=25)
        label.pack(side=LEFT)
        entry = ttk.Entry(frame, width=30, validate='key', validatecommand=self.vcmd) if validate else ttk.Entry(frame, width=30)
        entry.pack(side=LEFT, fill=X, expand=YES)
        ToolTip(entry, text=tooltip_text)
        return entry

    def validate_input(self, P):
        if P == "": return True
        if all(c in "0123456789.," for c in P) and P.count('.') <= 1 and P.count(',') <= 1:
            return True
        return False

    def format_currency(self, value):
        return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def parse_float(self, value_str):
        if not value_str: return 0.0
        return float(value_str.replace(",", "."))

    def copy_to_clipboard(self, text):
        self.clipboard_clear()
        numeric_value = text.replace("R$ ", "").strip()
        self.clipboard_append(numeric_value)
        Messagebox.show_info(f"Valor '{numeric_value}' copiado!", "Copiado", parent=self)

    def calcular_preco(self):
        try:
            self.inputs = self._get_input_data()
            
            total_percentual = self.inputs['margem_lucro_pct'] + self.inputs['custo_transacao_pct'] + self.inputs['impostos_pct']
            if total_percentual >= 100:
                Messagebox.show_error(
                    f"A soma das porcentagens (Lucro, Taxas, Impostos) não pode ser 100% ou mais. Soma atual: {total_percentual:.2f}%",
                    "Erro de Cálculo", parent=self
                )
                return

            custo_total = self.inputs['custo_produto'] + self.inputs['gasto_operacional']
            
            denominador = 1 - (total_percentual / 100)
            preco_venda_ideal = custo_total / denominador if denominador != 0 else 0
            
            imposto_valor = preco_venda_ideal * (self.inputs['impostos_pct'] / 100)
            custo_transacao_valor = preco_venda_ideal * (self.inputs['custo_transacao_pct'] / 100)
            lucro_estimado = preco_venda_ideal - custo_total - custo_transacao_valor - imposto_valor

            self.results = {
                'preco_venda_ideal': preco_venda_ideal,
                'lucro_estimado': lucro_estimado
            }

            self.preco_ideal_var.set(self.format_currency(preco_venda_ideal))
            self.lucro_estimado_var.set(self.format_currency(lucro_estimado))

            unidades = self.inputs['unidades']
            if unidades > 0:
                self.preco_unitario_var.set(self.format_currency(preco_venda_ideal / unidades))
                self.lucro_unitario_var.set(self.format_currency(lucro_estimado / unidades))
                self.unit_result_frame.pack(fill=X, pady=5)
            else:
                self.unit_result_frame.pack_forget()
            
            self.update_details_panel(custo_total, imposto_valor, custo_transacao_valor, lucro_estimado, preco_venda_ideal)

        except (ValueError, TypeError):
            Messagebox.show_error("Por favor, insira valores numéricos válidos.", "Erro de Entrada", parent=self)

    def update_details_panel(self, custo_total, imposto_valor, taxa_valor, lucro, preco_final):
        for widget in self.details_frame.winfo_children():
            widget.destroy()

        if preco_final == 0: return

        details_data = {
            "Custo Total (Produto + Op.)": custo_total,
            "Impostos": imposto_valor,
            "Taxa de Transação": taxa_valor,
            "Lucro Bruto": lucro
        }

        header_frame = ttk.Frame(self.details_frame)
        header_frame.pack(fill=X, padx=5, pady=(5,0))
        ttk.Label(header_frame, text="Componente", font=("Helvetica", 10, "bold")).pack(side=LEFT, expand=True, anchor='w')
        ttk.Label(header_frame, text="Valor (R$)", font=("Helvetica", 10, "bold"), width=15).pack(side=LEFT, anchor='e')
        ttk.Label(header_frame, text="%", font=("Helvetica", 10, "bold"), width=10).pack(side=LEFT, anchor='e')
        
        ttk.Separator(self.details_frame, orient=HORIZONTAL).pack(fill=X, pady=(2, 5))

        for label, value in details_data.items():
            percentage = (value / preco_final) * 100 if preco_final > 0 else 0
            row = ttk.Frame(self.details_frame)
            row.pack(fill=X, padx=5, pady=2)
            ttk.Label(row, text=label).pack(side=LEFT, expand=True, anchor='w')
            ttk.Label(row, text=f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."), width=15).pack(side=LEFT, anchor='e')
            ttk.Label(row, text=f"{percentage:.2f}%", width=10).pack(side=LEFT, anchor='e')

        self.details_frame.pack(fill=X, pady=10)

    def resetar_campos(self):
        self.entry_nome_produto.delete(0, END)
        self.entry_custo_produto.delete(0, END)
        self.entry_gasto_operacional.delete(0, END)
        self.entry_impostos_pct.delete(0, END)
        self.entry_margem_lucro.delete(0, END)
        self.entry_custo_transacao.delete(0, END)
        self.entry_unidades.delete(0, END)
        self.preco_ideal_var.set("R$ 0,00")
        self.lucro_estimado_var.set("R$ 0,00")
        self.preco_unitario_var.set("R$ 0,00")
        self.lucro_unitario_var.set("R$ 0,00")
        self.entry_nome_produto.focus()
        self.unit_result_frame.pack_forget()
        self.details_frame.pack_forget()
        self.inputs = {}
        self.results = {}

    def salvar_excel(self):
        if not self.results or self.results.get('preco_venda_ideal', 0) == 0:
            Messagebox.show_warning("Realize um cálculo antes de salvar.", "Aviso", parent=self)
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Planilhas Excel", "*.xlsx"), ("Todos os Arquivos", "*.*")],
            title="Salvar Histórico de Cálculo",
            initialfile="historico_de_calculos.xlsx"
        )

        if not filepath:
            return

        headers = [
            "Data e Hora", "Nome do Produto", "Custo do Produto (R$)", "Gasto Operacional (R$)",
            "Imposto (%)", "Taxa Transação (%)", "Margem Lucro (%)",
            "Preço de Venda Total (R$)", "Lucro Bruto Total (R$)"
        ]

        dados = [
            datetime.now().strftime("%d-%m-%Y %H:%M:%S"),
            self.inputs['nome_produto'],
            f"{self.inputs['custo_produto']:.2f}".replace(".", ","),
            f"{self.inputs['gasto_operacional']:.2f}".replace(".", ","),
            f"{self.inputs['impostos_pct']:.2f}".replace(".", ","),
            f"{self.inputs['custo_transacao_pct']:.2f}".replace(".", ","),
            f"{self.inputs['margem_lucro_pct']:.2f}".replace(".", ","),
            f"{self.results['preco_venda_ideal']:.2f}".replace(".", ","),
            f"{self.results['lucro_estimado']:.2f}".replace(".", ",")
        ]

        unidades = self.inputs.get('unidades', 0)
        if unidades > 0:
            headers.extend(["Unidades", "Preço por Unidade (R$)", "Lucro por Unidade (R$)"])
            preco_unitario = self.results['preco_venda_ideal'] / unidades
            lucro_unitario = self.results['lucro_estimado'] / unidades
            dados.extend([
                f"{unidades:.0f}",
                f"{preco_unitario:.2f}".replace(".", ","),
                f"{lucro_unitario:.2f}".replace(".", ",")
            ])

        try:
            is_new_file = not os.path.exists(filepath)
            workbook = openpyxl.load_workbook(filepath) if not is_new_file else openpyxl.Workbook()
            sheet = workbook.active
            
            if is_new_file or (sheet.max_row == 1 and sheet.cell(1,1).value is None):
                sheet.title = "Histórico"
                sheet.append(headers)
                for i, header in enumerate(headers, 1):
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = len(header) + 5
            
            sheet.append(dados)
            workbook.save(filepath)
            Messagebox.show_info(f"Dados salvos com sucesso em:\n{os.path.abspath(filepath)}", "Sucesso", parent=self)

        except PermissionError:
             Messagebox.show_error(f"O arquivo '{os.path.basename(filepath)}' está aberto ou você não tem permissão de escrita. Por favor, feche-o e tente novamente.", "Erro ao Salvar", parent=self)
        except Exception as e:
            Messagebox.show_error(f"Ocorreu um erro inesperado ao salvar o arquivo:\n{e}", "Erro ao Salvar", parent=self)

# --- FIM DA CLASSE DA CALCULADORA ---


# --- Classe Principal da Aplicação ---
class SalesApp:
    """Classe principal da aplicação de gestão de vendas."""
    TOTAL_ROW_ID = "total_row_id"

    def __init__(self, root):
        self.root = root
        self._setup_paths_and_dirs()
        self._setup_style()
        
        self.root.title("Sistema de Gestão de Vendas e Precificação")
        self.root.geometry("1150x850")

        self.db_manager = DatabaseManager(self.db_path)
        self.id_venda_em_edicao = None
        
        self.graph_canvas = {} # Dicionário para guardar os canvases dos gráficos

        self._create_widgets()
        self.atualizar_tabela()
        self.limpar_campos_e_resetar_edicao()
        self.root.protocol("WM_DELETE_WINDOW", self.fechar_app)

    def _setup_paths_and_dirs(self):
        """Define os caminhos para os arquivos e diretórios de dados."""
        try:
            app_dir = os.path.dirname(os.path.abspath(__file__))
            data_dir = os.path.join(app_dir, "_internal_data")
            os.makedirs(data_dir, exist_ok=True)
            if not os.access(data_dir, os.W_OK):
                raise OSError("Diretório de dados não é gravável.")
        except (OSError, NameError):
            data_dir = os.path.join(os.path.expanduser("~"), "Documents", "GestaoVendasData")
            os.makedirs(data_dir, exist_ok=True)
            messagebox.showwarning("Aviso de Caminho", f"Os dados serão salvos em: {data_dir}")

        self.db_path = os.path.join(data_dir, "vendas_gestao.db")
        self.theme_settings_path = os.path.join(data_dir, "theme_setting.txt")
        self.excel_export_folder = os.path.join(os.path.expanduser("~"), "Documents", "Vendas_Padaria_Exportadas")
        os.makedirs(self.excel_export_folder, exist_ok=True)

    def _setup_style(self):
        """Configura o tema ttkbootstrap."""
        self.current_theme_name = self._load_theme_setting()
        self.style = Style(theme=self.current_theme_name)
        self.style.configure("TButton", padding=(10, 5))

    def _create_widgets(self):
        """Cria e organiza todos os widgets da interface gráfica."""
        # --- Frame Superior (Título e Tema) ---
        top_frame = ttk.Frame(self.root, padding=(10, 10, 10, 0))
        top_frame.pack(fill=X)

        ttk.Label(top_frame, text="SISTEMA DE GESTÃO INTEGRADO", font=("Arial", 28, "bold")).pack(side=LEFT, expand=True)
        
        frame_theme = ttk.Frame(top_frame)
        frame_theme.pack(side=RIGHT)
        ttk.Label(frame_theme, text="Tema:").pack(side=tk.LEFT, padx=5)
        self.available_themes = sorted(list(self.style.theme_names()))
        self.theme_combobox = ttk.Combobox(frame_theme, values=self.available_themes, state="readonly", width=15)
        self.theme_combobox.set(self.current_theme_name)
        self.theme_combobox.pack(side=tk.LEFT)
        self.theme_combobox.bind("<<ComboboxSelected>>", self.change_theme)

        # --- Notebook (Sistema de Abas) ---
        notebook = ttk.Notebook(self.root, padding=10)
        notebook.pack(fill=BOTH, expand=True)

        # --- Aba 1: Gestão de Vendas ---
        tab_vendas = ttk.Frame(notebook)
        notebook.add(tab_vendas, text="Gestão de Vendas")
        self._create_vendas_tab(tab_vendas)

        # --- Aba 2: Calculadora de Preço (AGORA INTEGRADA) ---
        tab_calculadora = ttk.Frame(notebook)
        notebook.add(tab_calculadora, text="Calculadora de Preço")
        # Instancia a nova classe da calculadora, passando a aba como pai
        CalculadoraPrecoVenda(tab_calculadora, self.style).pack(fill=BOTH, expand=True)

        # --- Aba 3: Dashboard de Análise ---
        self.tab_dashboard = ttk.Frame(notebook, padding=10)
        notebook.add(self.tab_dashboard, text="Dashboard de Análise")
        self._create_dashboard_tab(self.tab_dashboard)

    def _create_vendas_tab(self, parent_tab):
        """Cria todos os widgets da aba de gestão de vendas. (APRIMORADO COM TOOLTIPS)"""
        parent_tab.grid_rowconfigure(2, weight=1)
        parent_tab.grid_columnconfigure(0, weight=1)

        # --- Formulário de Venda ---
        frame_formulario = ttk.LabelFrame(parent_tab, text="Registro de Venda", padding=10)
        frame_formulario.grid(row=0, column=0, padx=10, pady=10, sticky="ew")
        frame_formulario.grid_columnconfigure(1, weight=1)
        
        # Dicionário para Tooltips (APRIMORAMENTO)
        tooltips = {
            "Nome do Cliente:": "Nome do cliente ou 'Consumidor Final'.",
            "Nome do Produto:": "O item vendido (ex: Pão Francês, Bolo de Chocolate).",
            "Quantidade:": "Número de itens vendidos (somente números inteiros positivos).",
            "Valor por unidade (R$):": "Preço unitário. Use vírgula para centavos (ex: 5,50).",
            "Tipo de Pagamento:": "Selecione a forma de pagamento.",
            "Nome do Vendedor:": "Seu nome ou nome do funcionário.",
        }

        labels_info = ["Nome do Cliente:", "Nome do Produto:", "Quantidade:", "Valor por unidade (R$):", "Tipo de Pagamento:", "Nome do Vendedor:"]
        self.campos = {}
        for i, text in enumerate(labels_info):
            ttk.Label(frame_formulario, text=text).grid(row=i, column=0, sticky="w", padx=5, pady=5)
            if text == "Tipo de Pagamento:":
                widget = ttk.Combobox(frame_formulario, values=["Dinheiro", "Cartão de Crédito", "Cartão de Débito", "Pix"], state="readonly")
            else:
                widget = ttk.Entry(frame_formulario)
            widget.grid(row=i, column=1, sticky="ew", padx=5, pady=5)
            self.campos[text.split(':')[0]] = widget
            
            # Aplica o ToolTip (APRIMORAMENTO)
            ToolTip(widget, text=tooltips.get(text, "Campo de preenchimento."))


        frame_botoes_form = ttk.Frame(frame_formulario)
        frame_botoes_form.grid(row=len(labels_info), column=0, columnspan=2, pady=10, sticky="ew")
        self.btn_salvar = ttk.Button(frame_botoes_form, text="Registrar Venda", command=self.salvar_dados, style="success.TButton")
        self.btn_salvar.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        self.btn_cancelar_edicao = ttk.Button(frame_botoes_form, text="Cancelar Edição", command=self.limpar_campos_e_resetar_edicao, style="secondary.TButton")

        # --- Botões de Gestão ---
        frame_gestao = ttk.LabelFrame(parent_tab, text="Ações", padding=10)
        frame_gestao.grid(row=1, column=0, padx=10, pady=5, sticky="ew")
        
        botoes_gestao = {
            "Carregar para Editar": (self.carregar_para_edicao, "info.TButton"),
            "Excluir Selecionada(s)": (self.excluir_venda_selecionada, "danger.TButton"),
            "Exportar Vendas/Encomendas": (self.exportar_dados, "primary.TButton"),
            "Importar Planilha": (self.iniciar_importacao, "success.TButton"),
            "Abrir Planilha": (self.abrir_planilha_excel, "info.TButton"),
            "Calculadora": (lambda: subprocess.Popen('calc.exe') if os.name == 'nt' else subprocess.Popen('gnome-calculator'), "secondary.TButton"),
            "Encomendas": (self.abrir_caderno_encomendas, "success.TButton"),
            "Anotações": (self.abrir_anotacoes, "success.TButton"),
        }
        for i, (text, (command, style)) in enumerate(botoes_gestao.items()):
            ttk.Button(frame_gestao, text=text, command=command, style=style).pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
            frame_gestao.grid_columnconfigure(i, weight=1)

        # --- Tabela de Vendas ---
        frame_tabela = ttk.LabelFrame(parent_tab, text="Vendas Registradas", padding=10)
        frame_tabela.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")
        frame_tabela.grid_rowconfigure(1, weight=1)
        frame_tabela.grid_columnconfigure(0, weight=1)

        frame_pesquisa = ttk.Frame(frame_tabela, padding=(0, 5))
        frame_pesquisa.grid(row=0, column=0, sticky="ew")
        ttk.Label(frame_pesquisa, text="Pesquisar:").pack(side=tk.LEFT, padx=(0, 5))
        self.search_entry = ttk.Entry(frame_pesquisa)
        self.search_entry.pack(side=tk.LEFT, expand=True, fill=tk.X)
        self.search_entry.bind("<KeyRelease>", lambda e: self.atualizar_tabela())
        ttk.Button(frame_pesquisa, text="Limpar", command=lambda: [self.search_entry.delete(0, tk.END), self.atualizar_tabela()]).pack(side=tk.LEFT, padx=5)

        colunas = ("data_hora", "nome_cliente", "nome_produto", "quantidade", "preco", "tipo_pagamento", "preco_final", "nome_vendedor")
        self.tree = ttk.Treeview(frame_tabela, columns=colunas, show="headings", selectmode="extended")
        self.tree.grid(row=1, column=0, sticky="nsew")

        scrollbar_y = ttk.Scrollbar(frame_tabela, orient="vertical", command=self.tree.yview)
        scrollbar_y.grid(row=1, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scrollbar_y.set)

        headings = {"data_hora": "Data e Hora", "nome_cliente": "Cliente", "nome_produto": "Produto", "quantidade": "Qtd.",
                    "preco": "Valor Unid. (R$)", "tipo_pagamento": "Pagamento", "preco_final": "Preço Final (R$)", "nome_vendedor": "Vendedor"}
        widths = {"data_hora": 140, "quantidade": 60, "preco": 120, "tipo_pagamento": 120, "preco_final": 120}
        
        for col, text in headings.items():
            self.tree.heading(col, text=text)
            self.tree.column(col, width=widths.get(col, 150), anchor="center")

    def _create_dashboard_tab(self, parent_tab):
        """Cria os widgets da aba de Dashboard de Análise."""
        parent_tab.grid_rowconfigure(1, weight=1)
        parent_tab.grid_columnconfigure(0, weight=1)
        parent_tab.grid_columnconfigure(1, weight=1)

        # Frame de controle com botão de atualizar
        control_frame = ttk.Frame(parent_tab)
        control_frame.grid(row=0, column=0, columnspan=2, sticky="ew", padx=10, pady=(0,10))
        ttk.Button(control_frame, text="Atualizar Análise", command=self.update_dashboard, style="primary.TButton").pack(side=LEFT)

        # Frame principal para os gráficos e métricas
        main_dashboard_frame = ttk.Frame(parent_tab)
        main_dashboard_frame.grid(row=1, column=0, columnspan=2, sticky="nsew")
        main_dashboard_frame.grid_rowconfigure(1, weight=1) # Linha dos gráficos
        main_dashboard_frame.grid_columnconfigure(1, weight=2) # Coluna do gráfico maior

        # --- Métricas Principais ---
        metrics_frame = ttk.LabelFrame(main_dashboard_frame, text="Métricas Principais", padding=15)
        metrics_frame.grid(row=0, column=0, columnspan=2, sticky="ew", padx=10, pady=5)
        
        self.metric_labels = {
            "Faturamento Total": ttk.Label(metrics_frame, text="R$ 0,00", font=("Arial", 14, "bold")),
            "Total de Vendas": ttk.Label(metrics_frame, text="0", font=("Arial", 14, "bold")),
            "Ticket Médio": ttk.Label(metrics_frame, text="R$ 0,00", font=("Arial", 14, "bold")),
            "Produto Mais Vendido": ttk.Label(metrics_frame, text="-", font=("Arial", 14, "bold"), wraplength=250),
        }
        
        for i, (text, label_widget) in enumerate(self.metric_labels.items()):
            ttk.Label(metrics_frame, text=f"{text}:").grid(row=i, column=0, sticky="w", padx=5, pady=2)
            label_widget.grid(row=i, column=1, sticky="w", padx=5, pady=2)
        
        # --- Gráficos ---
        # Gráfico 1: Vendas ao longo do tempo (maior)
        graph_frame1 = ttk.LabelFrame(main_dashboard_frame, text="Faturamento por Período", padding=10)
        graph_frame1.grid(row=1, column=1, sticky="nsew", padx=10, pady=5)
        self.graph_canvas['vendas_tempo'] = self.create_placeholder_graph(graph_frame1)
        
        # Frame para os gráficos menores
        small_graphs_frame = ttk.Frame(main_dashboard_frame)
        small_graphs_frame.grid(row=1, column=0, sticky="ns", padx=10, pady=5)
        small_graphs_frame.grid_rowconfigure(0, weight=1)
        small_graphs_frame.grid_rowconfigure(1, weight=1)
        
        # Gráfico 2: Top 5 Produtos
        graph_frame2 = ttk.LabelFrame(small_graphs_frame, text="Top 5 Produtos (Quantidade)", padding=10)
        graph_frame2.grid(row=0, column=0, sticky="nsew", pady=(0, 5))
        self.graph_canvas['top_produtos'] = self.create_placeholder_graph(graph_frame2)

        # Gráfico 3: Formas de Pagamento
        graph_frame3 = ttk.LabelFrame(small_graphs_frame, text="Formas de Pagamento", padding=10)
        graph_frame3.grid(row=1, column=0, sticky="nsew", pady=(5, 0))
        self.graph_canvas['pagamentos'] = self.create_placeholder_graph(graph_frame3)
        
        # Inicia a aba com dados
        self.update_dashboard()

    def create_placeholder_graph(self, parent_frame):
        """Cria um canvas vazio para um gráfico."""
        fig = Figure(figsize=(5, 4), dpi=100)
        canvas = FigureCanvasTkAgg(fig, master=parent_frame)
        canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        return canvas

    def update_dashboard(self):
        """Busca os dados e atualiza todas as métricas e gráficos."""
        df_vendas = self.db_manager.fetch_sales_as_dataframe()
        
        if df_vendas.empty:
            # Limpa os gráficos se não houver dados
            for canvas in self.graph_canvas.values():
                canvas.figure.clear()
                canvas.draw()
            # Esta linha pode ser comentada para evitar pop-ups constantes
            # messagebox.showinfo("Dashboard", "Não há dados de vendas para analisar.", parent=self.root)
            return

        # Atualizar métricas
        faturamento_total = df_vendas['preco_final'].sum()
        total_vendas = len(df_vendas)
        ticket_medio = faturamento_total / total_vendas if total_vendas > 0 else 0
        produto_mais_vendido = df_vendas.groupby('nome_produto')['quantidade'].sum().idxmax() if not df_vendas.empty else "-"

        self.metric_labels["Faturamento Total"].config(text=f"R$ {faturamento_total:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        self.metric_labels["Total de Vendas"].config(text=f"{total_vendas}")
        self.metric_labels["Ticket Médio"].config(text=f"R$ {ticket_medio:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        self.metric_labels["Produto Mais Vendido"].config(text=produto_mais_vendido)

        # Atualizar gráficos
        self.plot_vendas_tempo(df_vendas, self.graph_canvas['vendas_tempo'])
        self.plot_top_produtos(df_vendas, self.graph_canvas['top_produtos'])
        self.plot_pagamentos(df_vendas, self.graph_canvas['pagamentos'])

    def _get_plot_style(self):
        """Retorna cores baseadas no tema atual para os gráficos."""
        bg_color = self.style.lookup('TFrame', 'background')
        fg_color = self.style.lookup('TLabel', 'foreground')
        return bg_color, fg_color

    def plot_vendas_tempo(self, df, canvas_widget):
        """Plota o gráfico de faturamento ao longo do tempo."""
        df_grouped = df.set_index('data_hora').resample('M')['preco_final'].sum()

        fig = canvas_widget.figure
        fig.clear()
        ax = fig.add_subplot(111)
        
        bg, fg = self._get_plot_style()
        fig.patch.set_facecolor(bg)
        ax.set_facecolor(bg)
        ax.tick_params(colors=fg, which='both')
        ax.spines['bottom'].set_color(fg)
        ax.spines['left'].set_color(fg)
        ax.spines['top'].set_color(bg)
        ax.spines['right'].set_color(bg)

        ax.plot(df_grouped.index, df_grouped.values, marker='o', linestyle='-', color=self.style.colors.primary)
        ax.set_title("Faturamento Mensal", color=fg)
        ax.set_xlabel("Mês", color=fg)
        ax.set_ylabel("Faturamento (R$)", color=fg)
        fig.autofmt_xdate()
        fig.tight_layout()
        canvas_widget.draw()

    def plot_top_produtos(self, df, canvas_widget):
        """Plota o gráfico de barras dos 5 produtos mais vendidos."""
        top_5 = df.groupby('nome_produto')['quantidade'].sum().nlargest(5)
        
        fig = canvas_widget.figure
        fig.clear()
        ax = fig.add_subplot(111)

        bg, fg = self._get_plot_style()
        fig.patch.set_facecolor(bg)
        ax.set_facecolor(bg)
        ax.tick_params(colors=fg, which='both')
        ax.spines['bottom'].set_color(fg)
        ax.spines['left'].set_color(fg)
        ax.spines['top'].set_color(bg)
        ax.spines['right'].set_color(bg)

        top_5.sort_values().plot(kind='barh', ax=ax, color=self.style.colors.info)
        ax.set_xlabel("Quantidade Vendida", color=fg)
        ax.set_ylabel("Produto", color=fg)
        fig.tight_layout()
        canvas_widget.draw()

    def plot_pagamentos(self, df, canvas_widget):
        """Plota o gráfico de pizza das formas de pagamento."""
        pagamentos = df['tipo_pagamento'].value_counts()
        
        fig = canvas_widget.figure
        fig.clear()
        ax = fig.add_subplot(111)

        bg, fg = self._get_plot_style()
        fig.patch.set_facecolor(bg)
        
        wedges, texts, autotexts = ax.pie(
            pagamentos, labels=pagamentos.index, autopct='%1.1f%%',
            startangle=90, textprops={'color': fg}
        )
        ax.axis('equal')
        fig.tight_layout()
        canvas_widget.draw()


    def abrir_caderno_encomendas(self):
        CadernoVirtual(self.root, self.style, self.db_manager)

    def abrir_anotacoes(self):
        AnotacoesVirtual(self.root, self.style, self.db_manager)

    def abrir_planilha_excel(self):
        # O nome do arquivo inicial precisa ser ajustado após a correção
        file_path = os.path.join(self.excel_export_folder, "vendas_gestao_exportadas.xlsx")
        if not os.path.exists(file_path):
            messagebox.showwarning("Atenção", "O arquivo Excel ainda não foi exportado. Por favor, exporte-o primeiro.")
            return
        
        try:
            if os.name == 'nt': subprocess.Popen(['start', '', file_path], shell=True)
            elif os.uname().sysname == 'Darwin': subprocess.Popen(['open', file_path])
            else: subprocess.Popen(['xdg-open', file_path])
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível abrir a planilha: {e}")

    def salvar_dados(self):
        try:
            dados = {
                "nome_cliente": self.campos["Nome do Cliente"].get().strip(),
                "nome_produto": self.campos["Nome do Produto"].get().strip(),
                "quantidade": self.campos["Quantidade"].get().strip(),
                "preco": self.campos["Valor por unidade (R$)"].get().strip(),
                "tipo_pagamento": self.campos["Tipo de Pagamento"].get().strip(),
                "nome_vendedor": self.campos["Nome do Vendedor"].get().strip(),
            }

            if not all(dados.values()):
                messagebox.showwarning("Atenção", "Todos os campos são obrigatórios.")
                return

            quantidade = int(dados["quantidade"])
            preco_unitario = float(dados["preco"].replace(",", "."))
            if quantidade <= 0 or preco_unitario < 0:
                raise ValueError("Valores devem ser positivos.")

            dados.update({
                "quantidade": quantidade,
                "preco": preco_unitario,
                "preco_final": quantidade * preco_unitario,
                "data_hora": datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            })

            if self.id_venda_em_edicao is not None:
                self.db_manager.update_sale(self.id_venda_em_edicao, dados)
                messagebox.showinfo("Sucesso", "Venda atualizada!")
            else:
                self.db_manager.insert_sale(dados)
                messagebox.showinfo("Sucesso", "Venda registrada!")

            self.limpar_campos_e_resetar_edicao()
            self.atualizar_tabela()

        except ValueError:
            messagebox.showerror("Erro de Validação", "Quantidade ou Preço inválidos. Verifique os valores inseridos.")
        except Exception as e:
            messagebox.showerror("Erro Inesperado", f"Ocorreu um erro: {e}")

    def limpar_campos_e_resetar_edicao(self):
        self.id_venda_em_edicao = None
        for widget in self.campos.values():
            if isinstance(widget, ttk.Combobox):
                widget.set("")
            else:
                widget.delete(0, tk.END)
        self.btn_salvar.config(text="Registrar Venda", style="success.TButton")
        self.btn_cancelar_edicao.pack_forget()

    def carregar_para_edicao(self):
        selecionado = self.tree.selection()
        if not selecionado or len(selecionado) > 1 or selecionado[0] == self.TOTAL_ROW_ID:
            messagebox.showwarning("Atenção", "Selecione uma única venda para editar.")
            return

        venda_id = int(selecionado[0])
        venda = self.db_manager.fetch_sale_by_id(venda_id)
        if venda:
            self.limpar_campos_e_resetar_edicao()
            _, nome_cli, nome_prod, qtd, preco, tipo_pag, nome_vend = venda
            
            self.campos["Nome do Cliente"].insert(0, nome_cli)
            self.campos["Nome do Produto"].insert(0, nome_prod)
            self.campos["Quantidade"].insert(0, str(qtd))
            self.campos["Valor por unidade (R$)"].insert(0, str(preco).replace(".", ","))
            self.campos["Tipo de Pagamento"].set(tipo_pag)
            self.campos["Nome do Vendedor"].insert(0, nome_vend)

            self.id_venda_em_edicao = venda_id
            self.btn_salvar.config(text="Atualizar Venda", style="primary.TButton")
            self.btn_cancelar_edicao.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)

    def excluir_venda_selecionada(self):
        selecionados = [item for item in self.tree.selection() if item != self.TOTAL_ROW_ID]
        if not selecionados:
            messagebox.showwarning("Atenção", "Selecione uma ou mais vendas para excluir.")
            return

        msg = f"Tem certeza que deseja excluir {len(selecionados)} venda(s)?"
        if messagebox.askyesno("Confirmar Exclusão", msg):
            for item_id in selecionados:
                self.db_manager.delete_sale(int(item_id))
            messagebox.showinfo("Sucesso", f"{len(selecionados)} venda(s) excluída(s).")
            self.atualizar_tabela()
            self.limpar_campos_e_resetar_edicao()

    # --- FUNÇÃO CORRIGIDA E MELHORADA PARA ESCOLHA DO LOCAL ---
    def exportar_dados(self):
        """Exporta todos os dados de Vendas e Encomendas para um único arquivo Excel, permitindo ao usuário escolher o local."""
        
        # 1. Confirmação
        if not messagebox.askyesno("Confirmar", "Deseja exportar todos os dados de Vendas e Encomendas?"):
            return

        # 2. CAIXA DE DIÁLOGO PARA ESCOLHER O CAMINHO DE SALVAMENTO
        excel_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Planilha Excel", "*.xlsx")],
            initialdir=self.excel_export_folder, # Sugere a pasta padrão
            initialfile="vendas_gestao_exportadas.xlsx",
            title="Escolha onde salvar a planilha de exportação"
        )
        
        if not excel_path:
            return # Usuário cancelou
        
        try:
            # 3. Obter DataFrames
            df_vendas = pd.read_sql_query("SELECT id, data_hora, nome_cliente, nome_produto, quantidade, preco, tipo_pagamento, preco_final, nome_vendedor FROM vendas", self.db_manager.conn)
            df_enc = pd.read_sql_query("SELECT id, data_hora_registro, nome_cliente, produto, quantidade, valor_unitario, data_entrega FROM encomendas", self.db_manager.conn)
            
            if df_vendas.empty and df_enc.empty:
                 messagebox.showwarning("Aviso", "Não há dados de Vendas nem Encomendas para exportar.")
                 return
                 
            # 4. Preparar DataFrames para exportação
            if not df_vendas.empty:
                df_vendas.rename(columns={'id': 'ID Venda', 'preco': 'Valor Unid. (R$)', 'preco_final': 'Preço Final (R$)'}, inplace=True)
            
            if not df_enc.empty:
                df_enc['Valor Total (R$)'] = df_enc['quantidade'] * df_enc['valor_unitario']
                df_enc.rename(columns={'id': 'ID Encomenda', 'valor_unitario': 'Valor Unid. (R$)'}, inplace=True)
            
            # 5. Exportar usando ExcelWriter
            with pd.ExcelWriter(excel_path, engine='xlsxwriter') as writer:
                
                workbook = writer.book
                # Formato monetário brasileiro
                money_format = workbook.add_format({'num_format': 'R$ #,##0.00'})
                
                # --- Exportar Vendas ---
                if not df_vendas.empty:
                    df_vendas.to_excel(writer, sheet_name='Vendas', index=False)
                    worksheet_vendas = writer.sheets['Vendas']
                    
                    # Aplica a formatação de moeda e largura
                    worksheet_vendas.set_column('F:F', 18, money_format) # Coluna de Valor Unid. (R$)
                    worksheet_vendas.set_column('H:H', 18, money_format) # Coluna de Preço Final (R$)
                    
                    # Ajusta a largura das colunas baseada nos cabeçalhos
                    for i, col in enumerate(df_vendas.columns):
                        worksheet_vendas.set_column(i, i, max(len(col) + 2, 12))


                # --- Exportar Encomendas ---
                if not df_enc.empty:
                    df_enc.to_excel(writer, sheet_name='Encomendas', index=False)
                    worksheet_enc = writer.sheets['Encomendas']
                    
                    # Aplica a formatação de moeda e largura
                    worksheet_enc.set_column('F:F', 18, money_format) # Valor Unid. (R$)
                    worksheet_enc.set_column('G:G', 18, money_format) # Valor Total (R$)
                    
                    # Ajusta a largura das colunas baseada nos cabeçalhos
                    for i, col in enumerate(df_enc.columns):
                        worksheet_enc.set_column(i, i, max(len(col) + 2, 12))
            
            # O bloco 'with' garante que o arquivo seja salvo e fechado corretamente.
            messagebox.showinfo("Sucesso", f"Dados exportados para:\n{excel_path}")
            
        except PermissionError:
            messagebox.showerror("Erro de Permissão", f"O arquivo '{os.path.basename(excel_path)}' está aberto ou você não tem permissão. Por favor, feche-o e tente novamente.")
        except Exception as e:
            messagebox.showerror("Erro de Exportação", f"Não foi possível exportar os dados:\n{e}")

    def iniciar_importacao(self):
        """Abre o diálogo de arquivo e inicia o processo de importação inteligente."""
        file_path = filedialog.askopenfilename(
            title="Selecione a planilha para importar",
            filetypes=[("Planilhas Excel", "*.xlsx"), ("Arquivos CSV", "*.csv")]
        )
        if not file_path:
            return

        try:
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path, keep_default_na=False) # keep_default_na para não converter "NA" em NaN
            else:
                df = pd.read_excel(file_path, keep_default_na=False)

            # Tenta o mapeamento automático primeiro
            auto_mapping, missing_fields = self.tentar_mapeamento_automatico(df.columns)
            
            if not missing_fields: # Se todos os campos obrigatórios foram encontrados
                if messagebox.askyesno("Mapeamento Automático", "O sistema identificou as colunas necessárias. Deseja prosseguir com a importação?"):
                    self.processar_importacao(df, auto_mapping)
                else: # Se o usuário quiser revisar
                    ImportMappingWindow(self, df, self.processar_importacao, initial_mapping=auto_mapping)
            else: # Se faltarem campos, abre a janela de mapeamento manual
                messagebox.showwarning("Mapeamento Incompleto", f"Não foi possível encontrar colunas para: {', '.join(missing_fields)}. Por favor, mapeie manualmente.")
                ImportMappingWindow(self, df, self.processar_importacao, initial_mapping=auto_mapping)

        except Exception as e:
            messagebox.showerror("Erro ao Ler Arquivo", f"Não foi possível ler a planilha:\n{e}")

    def tentar_mapeamento_automatico(self, sheet_columns):
        """Tenta mapear automaticamente as colunas da planilha para os campos do sistema."""
        required_fields = {
            "Nome do Cliente", "Nome do Produto", "Quantidade", "Valor por unidade (R$)"
        }
        
        mapping = {}
        found_fields = set()

        # Primeiro, tenta encontrar os mapeamentos
        for field in self.system_fields:
            for col in sheet_columns:
                if self.guess_mapping(col) == field and col not in mapping:
                    mapping[col] = field
                    if field in required_fields:
                        found_fields.add(field)
                    break # Vai para o próximo campo do sistema

        missing_fields = required_fields - found_fields
        return mapping, missing_fields

    def processar_importacao(self, df, mapping):
        """Processa o DataFrame mapeado e insere no banco de dados."""
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Importando Dados")
        progress_window.geometry("300x100")
        progress_window.transient(self.root)
        progress_window.grab_set()

        ttk.Label(progress_window, text="Processando linhas...").pack(pady=10)
        progress_bar = ttk.Progressbar(progress_window, length=250, mode='determinate', maximum=len(df))
        progress_bar.pack(pady=5)

        threading.Thread(target=self._worker_import, args=(df, mapping, progress_bar, progress_window), daemon=True).start()

    def _worker_import(self, df, mapping, progress_bar, progress_window):
        """(Worker Thread) Converte e insere os dados da planilha."""
        sales_to_insert = []
        system_to_sheet = {v: k for k, v in mapping.items()}

        for index, row in df.iterrows():
            try:
                nome_cliente = str(row[system_to_sheet["Nome do Cliente"]])
                nome_produto = str(row[system_to_sheet["Nome do Produto"]])
                quantidade = int(row[system_to_sheet["Quantidade"]])
                preco_unitario = float(str(row[system_to_sheet["Valor por unidade (R$)"]]).replace(",", "."))
                
                tipo_pagamento = str(row.get(system_to_sheet.get("Tipo de Pagamento"), "Não Informado"))
                nome_vendedor = str(row.get(system_to_sheet.get("Nome do Vendedor"), "Importado"))
                
                if not nome_cliente or not nome_produto or quantidade <= 0 or preco_unitario < 0:
                    continue

                preco_final = quantidade * preco_unitario
                data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

                sales_to_insert.append((
                    nome_cliente, nome_produto, quantidade, preco_unitario, tipo_pagamento,
                    preco_final, nome_vendedor, data_hora
                ))
            except (KeyError, ValueError, TypeError) as e:
                print(f"Aviso: Pulando linha {index+2} da planilha devido a dados inválidos. Erro: {e}")
                continue

            self.root.after(0, lambda p=index+1: progress_bar.config(value=p))
        
        if sales_to_insert:
            success = self.db_manager.insert_multiple_sales(sales_to_insert)
            if success:
                self.root.after(0, lambda: messagebox.showinfo("Sucesso", f"{len(sales_to_insert)} registros importados com sucesso!"))
        else:
            self.root.after(0, lambda: messagebox.showwarning("Aviso", "Nenhum registro válido encontrado para importação."))

        self.root.after(0, progress_window.destroy)
        self.root.after(0, self.atualizar_tabela)
    
    @property
    def system_fields(self):
        """Retorna os campos do sistema que podem ser mapeados."""
        return [
            "Nome do Cliente", "Nome do Produto", "Quantidade", 
            "Valor por unidade (R$)", "Tipo de Pagamento", "Nome do Vendedor"
        ]

    def guess_mapping(self, column_name):
        """Tenta adivinhar o campo do sistema com mais variações, em ordem de especificidade."""
        col = column_name.lower().strip()
        # Mais específico primeiro
        if any(term in col for term in ["cliente", "comprador"]): return "Nome do Cliente"
        if any(term in col for term in ["produto", "item", "descrição"]): return "Nome do Produto"
        if any(term in col for term in ["qtd", "quant", "quantidade"]): return "Quantidade"
        if any(term in col for term in ["valor unit", "preço unit", "preco unit", "valor unid", "preço unid"]): return "Valor por unidade (R$)"
        # Termos mais genéricos depois
        if "vendedor" in col: return "Nome do Vendedor"
        if any(term in col for term in ["valor", "preço", "preco"]): return "Valor por unidade (R$)"
        if "pagamento" in col or any(term in col for term in ["forma", "tipo"]): return "Tipo de Pagamento"
        if "nome" in col: return "Nome do Cliente" # Menos específico, por último
        return "Ignorar"

    def change_theme(self, event=None):
        selected_theme = self.theme_combobox.get()
        self.style.theme_use(selected_theme)
        self.current_theme_name = selected_theme
        self._save_theme_setting(selected_theme)
        self.atualizar_tabela()
        self.update_dashboard()

    def _load_theme_setting(self):
        try:
            with open(self.theme_settings_path, 'r') as f:
                theme = f.read().strip()
            return theme if theme else "cosmo"
        except FileNotFoundError:
            return "cosmo"

    def _save_theme_setting(self, theme_name):
        with open(self.theme_settings_path, 'w') as f:
            f.write(theme_name)

    def atualizar_tabela(self):
        self.tree.delete(*self.tree.get_children())
        search_term = self.search_entry.get().strip()
        vendas = self.db_manager.fetch_all_sales(search_term)
        
        total_geral = 0.0
        for venda in vendas:
            venda_id, data_h, nome_c, nome_p, qtd, preco_u, tipo_p, preco_f, nome_v = venda
            self.tree.insert('', tk.END, iid=venda_id, values=(
                data_h, nome_c, nome_p, qtd,
                f"{preco_u:.2f}".replace(".", ","),
                tipo_p,
                f"{preco_f:.2f}".replace(".", ","),
                nome_v
            ))
            total_geral += preco_f

        self.tree.insert('', tk.END, iid=self.TOTAL_ROW_ID, values=(
            "", "", "", "", "", "", f"TOTAL: {total_geral:.2f}".replace(".",","), ""
        ), tags=('total_row',))
        self.tree.tag_configure('total_row', background=self.style.lookup('TFrame', 'background'), font=('Arial', 10, 'bold'))

    def fechar_app(self):
        self.db_manager.close_connection()
        self.root.destroy()

# --- Janela de Mapeamento de Importação ---
class ImportMappingWindow(tk.Toplevel):
    def __init__(self, parent_app, dataframe, callback, initial_mapping=None): # CORREÇÃO: Recebe a instância da app
        super().__init__(parent_app.root)
        self.parent_app = parent_app # CORREÇÃO: Armazena a referência
        self.title("Mapeamento de Importação")
        self.geometry("600x400")
        self.transient(parent_app.root)
        self.grab_set()

        self.df = dataframe
        self.callback = callback
        self.column_mappings = {}
        self.system_fields = self.parent_app.system_fields

        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill=BOTH, expand=True)
        
        ttk.Label(main_frame, text="Associe as colunas da sua planilha com os campos do sistema:", wraplength=550).pack(pady=(0,10))

        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=X, pady=5)
        ttk.Label(header_frame, text="Coluna da Planilha", font=("Arial", 10, "bold")).pack(side=LEFT, expand=True)
        ttk.Label(header_frame, text="Campo do Sistema", font=("Arial", 10, "bold")).pack(side=RIGHT, expand=True)

        scroll_canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=scroll_canvas.yview)
        mapping_frame = ttk.Frame(scroll_canvas)

        scroll_canvas.configure(yscrollcommand=scrollbar.set)
        scroll_canvas.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.pack(side=RIGHT, fill=Y)
        scroll_canvas.create_window((0, 0), window=mapping_frame, anchor="nw")
        mapping_frame.bind("<Configure>", lambda e: scroll_canvas.configure(scrollregion=scroll_canvas.bbox("all")))

        for col in self.df.columns:
            frame = ttk.Frame(mapping_frame, padding=5)
            frame.pack(fill=X, expand=True)
            ttk.Label(frame, text=col).pack(side=LEFT)
            
            combo = ttk.Combobox(frame, values=["Ignorar"] + self.system_fields, state="readonly")
            
            # Usa o mapeamento inicial se fornecido
            if initial_mapping and col in initial_mapping:
                combo.set(initial_mapping[col])
            else:
                combo.set(self.parent_app.guess_mapping(col)) # CORREÇÃO: Chama o método da instância da app

            combo.pack(side=RIGHT)
            self.column_mappings[col] = combo

        button_frame = ttk.Frame(self, padding=10)
        button_frame.pack(fill=X)
        ttk.Button(button_frame, text="Confirmar Importação", command=self.confirm, style="success.TButton").pack(side=LEFT, expand=True)
        ttk.Button(button_frame, text="Cancelar", command=self.destroy, style="secondary.TButton").pack(side=RIGHT, expand=True)

    def confirm(self):
        final_mapping = {col: combo.get() for col, combo in self.column_mappings.items() if combo.get() != "Ignorar"}
        
        required_fields = {"Nome do Cliente", "Nome do Produto", "Quantidade", "Valor por unidade (R$)"}
        mapped_fields = set(final_mapping.values())

        if not required_fields.issubset(mapped_fields):
            messagebox.showerror("Mapeamento Incompleto", "Os campos obrigatórios (Cliente, Produto, Quantidade, Valor) devem ser mapeados.", parent=self)
            return
            
        self.destroy()
        self.callback(self.df, final_mapping)


if __name__ == "__main__":
    root = tk.Tk()
    
    # --- Definir o ícone da janela (Método aprimorado para Windows) ---
    try:
        icon_path_abs = resource_path("cesto.ico") 
        root.iconbitmap(icon_path_abs) 

        # Método avançado para forçar o ícone na barra de tarefas (apenas para Windows)
        if os.name == 'nt':
            # Força a criação do "handle" da janela no Windows
            root.update()
            
            # Obtém o "handle" (identificador) da janela
            hwnd = ctypes.windll.user32.GetParent(root.winfo_id())
            
            # Constantes da API do Windows para definir o ícone
            ICON_BIG = 1
            WM_SETICON = 0x0080

            # Envia uma mensagem direta para o Windows, definindo o ícone da janela
            ctypes.windll.user32.SendMessageW(hwnd, WM_SETICON, ICON_BIG, ctypes.windll.user32.LoadImageW(None, icon_path_abs, 1, 0, 0, 0x0010))

    except Exception as e:
        print(f"AVISO: Não foi possível carregar o ícone da janela. Erro: {e}")
        
    app = SalesApp(root)
    root.mainloop()